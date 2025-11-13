"""Utilities for exporting report data to CSV/XLSX."""

from __future__ import annotations

from io import BytesIO
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _format_number(value):
    if value is None:
        return ''
    if isinstance(value, float):
        return f'{value:.2f}'
    return value


def _write_table(ws, headers: Iterable[str], rows: Iterable[Iterable]):
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        # hier goenne ich den Spalten eine fixe Mindestbreite, damit niemand nach dem Export nachjustieren muss
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 14), 40)


def build_report_xlsx(report: Dict, temperature_samples: List[Dict],
                      ui_strings: Dict[str, str]) -> bytes:
    """Return XLSX bytes for the supplied report."""
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = ui_strings.get('report_table_period', 'Period')[:31] or 'Summary'

    params = report['params']
    # die wichtigsten Filter schreibe ich bewusst nochmal in die erste Tabelle rein
    summary_rows = [
        (ui_strings.get('report_table_period', 'Period'),
         params.get('start_date'), params.get('end_date')),
        (ui_strings.get('radius_label', 'Radius'),
         f"{params.get('radius', 0):.0f} km", ''),
        (ui_strings.get('report_station_usage_found', 'Stations found'),
         report.get('station_count', 0),
         f"{ui_strings.get('report_station_usage_used', 'Used')}: {report.get('used_station_count', 0)}"),
    ]
    for row in summary_rows:
        summary_ws.append(row)
    for row in summary_ws.iter_rows(min_row=1, max_col=3, max_row=len(summary_rows)):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
            cell.font = Font(bold=True) if cell.column == 1 else Font()
    summary_ws.column_dimensions['A'].width = 28
    summary_ws.column_dimensions['B'].width = 22
    summary_ws.column_dimensions['C'].width = 22

    periods_ws = wb.create_sheet(ui_strings.get('report_table_title', 'Table')[:31] or 'Periods')
    _write_table(
        periods_ws,
        [
            ui_strings.get('report_table_period', 'Period'),
            ui_strings.get('report_table_temp_avg', 'Ø Temp'),
            ui_strings.get('report_table_temp_min', 'Min Temp'),
            ui_strings.get('report_table_temp_max', 'Max Temp'),
            ui_strings.get('report_table_precip', 'Precipitation'),
            ui_strings.get('report_table_sunshine', 'Sunshine'),
            ui_strings.get('report_period_details_heading', 'Stations'),
        ],
        (
            (
                row.get('period'),
                row.get('temp_avg'),
                row.get('temp_min'),
                row.get('temp_max'),
                row.get('precipitation'),
                row.get('sunshine'),
                len(row.get('stations') or []),
            )
            for row in report['periods']
        ),
    )

    stations_ws = wb.create_sheet(ui_strings.get('report_stations_heading', 'Stations')[:31] or 'Stations')
    _write_table(
        stations_ws,
        [
            ui_strings.get('report_stations_heading', 'Station'),
            ui_strings.get('temp_table_distance', 'Distance'),
            ui_strings.get('report_station_usage_used', 'Used'),
        ],
        (
            (
                f"{station.get('name') or station.get('station_id')} (ID {station.get('station_id')})"
                + (f" – {station.get('state')}" if station.get('state') else ''),
                station.get('distance_km'),
                'Yes' if station.get('has_data') else 'No',
            )
            for station in report['stations']
        ),
    )

    if temperature_samples:
        samples_ws = wb.create_sheet(ui_strings.get('temp_table_heading', 'Samples')[:31] or 'Samples')
        _write_table(
            samples_ws,
            [
                ui_strings.get('temp_table_date', 'Date'),
                ui_strings.get('temp_table_station', 'Station'),
                ui_strings.get('temp_table_distance', 'Distance'),
                ui_strings.get('temp_table_temperature', 'Temperature'),
            ],
            (
                (
                    sample.get('date'),
                    f"{sample.get('station_name') or sample.get('station_id')} (ID {sample.get('station_id')})"
                    + (f" – {sample.get('state')}" if sample.get('state') else ''),
                    sample.get('distance_km'),
                    sample.get('temperature'),
                )
                for sample in temperature_samples
            ),
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
